import openpyxl
import re

wb = openpyxl.load_workbook(r'c:\laragon\www\eSPPD\md\DatabaseDosen.xlsx')
ws = wb.active

print('=== ANALYZING EXCEL COLUMNS ===\n')
print('ROW 2 DATA:\n')
for col in range(1, 14):
    cell_value = ws.cell(2, col).value
    if isinstance(cell_value, str) and 'IFERROR' in cell_value:
        match = re.search(r'"([^"]+)"\s*\)', cell_value)
        if match:
            extracted = match.group(1)
            # Remove invisible characters
            cleaned = ''.join(c for c in extracted if ord(c) >= 32)
            print(f'Col {col:2d}: "{cleaned[:40]}" (len={len(cleaned)})')
        else:
            print(f'Col {col:2d}: [FORMULA - no match]')
    else:
        print(f'Col {col:2d}: {str(cell_value)[:40]}')

# Test: check row 3 too
print('\n\nROW 3 DATA:\n')
for col in range(1, 8):
    cell_value = ws.cell(3, col).value
    if isinstance(cell_value, str) and 'IFERROR' in cell_value:
        match = re.search(r'"([^"]+)"\s*\)', cell_value)
        if match:
            extracted = match.group(1)
            cleaned = ''.join(c for c in extracted if ord(c) >= 32)
            print(f'Col {col:2d}: "{cleaned[:40]}" (len={len(cleaned)})')
