#!/usr/bin/env python3
"""
Login Credentials Generator for e-SPPD
Exports NIP + DDMMYYYY format credentials from database to Excel

Usage:
    python export_credentials.py --output storage/credentials.xlsx
"""

import argparse
import os
import sys
import io
from datetime import datetime
from pathlib import Path

# Fix Windows encoding issues
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')

try:
    import pandas as pd
    import psycopg2
    from dotenv import load_dotenv
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
except ImportError as e:
    print(f"Missing dependency: {e}")
    print("Install with: pip install pandas psycopg2-binary python-dotenv openpyxl")
    sys.exit(1)


class CredentialsExporter:
    """Export login credentials from database to Excel"""

    def __init__(self, db_config: dict):
        self.db_config = db_config
        self.conn = None
        self.users = []

    def connect(self):
        """Connect to PostgreSQL database"""
        try:
            self.conn = psycopg2.connect(
                host=self.db_config['host'],
                port=self.db_config['port'],
                database=self.db_config['database'],
                user=self.db_config['user'],
                password=self.db_config['password']
            )
            print(f"✅ Connected to database: {self.db_config['database']}")
        except psycopg2.Error as e:
            print(f"❌ Connection failed: {e}")
            sys.exit(1)

    def fetch_users(self):
        """Fetch all users with birth dates from database"""
        try:
            cursor = self.conn.cursor()

            # Query users with LEFT JOIN to employees for birth_date
            query = """
                SELECT
                    u.id,
                    u.nip,
                    u.name,
                    u.email,
                    e.birth_date,
                    u.created_at
                FROM users u
                LEFT JOIN employees e ON u.id = e.user_id
                WHERE u.nip IS NOT NULL AND u.nip != ''
                ORDER BY u.nip
            """

            cursor.execute(query)
            rows = cursor.fetchall()
            cursor.close()

            print(f"✅ Fetched {len(rows)} users from database")

            # Convert to list of dicts
            for row in rows:
                self.users.append({
                    'id': row[0],
                    'nip': row[1].strip() if row[1] else '',
                    'name': row[2],
                    'email': row[3],
                    'birth_date': row[4],
                    'created_at': row[5]
                })

            return len(self.users)

        except psycopg2.Error as e:
            print(f"❌ Query failed: {e}")
            sys.exit(1)

    def generate_password(self, birth_date) -> tuple:
        """
        Generate password from birth date (DDMMYYYY)
        Returns (plain_password, status_message)
        """
        if not birth_date:
            return ('PENDING', 'No birth date')

        try:
            # Handle datetime object
            if isinstance(birth_date, str):
                birth_date = datetime.strptime(birth_date, '%Y-%m-%d').date()

            # Format as DDMMYYYY
            plain_password = birth_date.strftime('%d%m%Y')
            return (plain_password, 'Complete')

        except Exception as e:
            return ('ERROR', f'Invalid date: {str(e)}')

    def export_to_excel(self, output_path: str):
        """Export credentials to Excel file"""
        print(f"\n📊 Exporting to Excel...")

        wb = Workbook()
        ws = wb.active
        ws.title = "Login Credentials"

        # Set headers
        headers = ['No', 'NIP', 'Nama Lengkap', 'Email', 'Tanggal Lahir', 'Password (NIP+DDMMYYYY)', 'Status']
        ws.append(headers)

        # Style header
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Add data
        count = 0
        with_birth = 0
        without_birth = 0

        print("Processing credentials...")
        progress_interval = max(1, len(self.users) // 10)

        for idx, user in enumerate(self.users):
            count += 1

            if (idx + 1) % progress_interval == 0:
                print(f"  {idx + 1}/{len(self.users)} users processed")

            # Generate password
            password, status = self.generate_password(user['birth_date'])

            if status == 'Complete':
                with_birth += 1
            else:
                without_birth += 1

            # Format birth date for display
            birth_date_str = ''
            if user['birth_date']:
                if isinstance(user['birth_date'], str):
                    birth_date_str = user['birth_date']
                else:
                    birth_date_str = user['birth_date'].strftime('%Y-%m-%d')

            # Add row
            ws.append([
                count,
                user['nip'],
                user['name'],
                user['email'] or '-',
                birth_date_str or '-',
                password,
                status
            ])

        # Auto-size columns
        ws.column_dimensions['A'].width = 6
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 35
        ws.column_dimensions['D'].width = 30
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 30
        ws.column_dimensions['G'].width = 20

        # Save file
        wb.save(output_path)

        print(f"\n✅ Export Complete!")
        print("=" * 70)
        print(f"File: {Path(output_path).name}")
        print(f"Location: {output_path}")
        print("=" * 70)

        print(f"\n📊 SUMMARY")
        print("━" * 70)
        print(f"Total users exported:       {count}")
        print(f"Users dengan birth date:    {with_birth}")
        print(f"Users tanpa birth date:     {without_birth}")

        if without_birth > 0:
            print(f"\n⚠️  Perhatian: {without_birth} user(s) belum punya tanggal lahir")
            print(f"   Status password: PENDING")
            print(f"   Diperlukan import data dari file Excel untuk melengkapi")

        print(f"\n✅ File siap digunakan untuk login credentials!")

        return {
            'file': output_path,
            'total': count,
            'with_birth': with_birth,
            'without_birth': without_birth
        }

    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()


def main():
    # Load environment variables
    load_dotenv()

    parser = argparse.ArgumentParser(description="Export login credentials from e-SPPD database")
    parser.add_argument("--output", "-o", default="LOGIN_CREDENTIALS.xlsx",
                       help="Output Excel file path (default: LOGIN_CREDENTIALS.xlsx)")
    parser.add_argument("--host", default=os.getenv("DB_HOST", "127.0.0.1"),
                       help="Database host")
    parser.add_argument("--port", type=int, default=int(os.getenv("DB_PORT", "5432")),
                       help="Database port")
    parser.add_argument("--database", default=os.getenv("DB_DATABASE", "esppd"),
                       help="Database name")
    parser.add_argument("--user", default=os.getenv("DB_USERNAME", "postgres"),
                       help="Database user")
    parser.add_argument("--password", default=os.getenv("DB_PASSWORD", ""),
                       help="Database password")

    args = parser.parse_args()

    db_config = {
        'host': args.host,
        'port': args.port,
        'database': args.database,
        'user': args.user,
        'password': args.password
    }

    print("====== LOGIN CREDENTIALS EXPORTER ======\n")

    # Export
    exporter = CredentialsExporter(db_config)
    try:
        exporter.connect()
        exporter.fetch_users()
        result = exporter.export_to_excel(args.output)
        return 0
    except Exception as e:
        print(f"❌ Error: {e}")
        return 1
    finally:
        exporter.close()


if __name__ == "__main__":
    sys.exit(main())
